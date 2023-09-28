import os
import json
from openpyxl import load_workbook

WORKING_DIR = os.path.dirname(os.path.realpath(__file__))
IN_FILE = os.path.join(WORKING_DIR, "Strings.xlsx")
OUT_FILE = os.path.join(WORKING_DIR, "TheOtherRoles", "Resources", "stringData.json")

def getLanguageId(str):
  match str:
    case "English":
      return 0
    case "Latam":
      return 1
    case "Brazilian":
      return 2
    case "Portuguese":
      return 3
    case "Korean":
      return 4
    case "Russian":
      return 5
    case "Dutch":
      return 6
    case "Filipino":
      return 7
    case "French":
      return 8
    case "German":
      return 9
    case "Italian":
      return 10
    case "Japanese":
      return 11
    case "Spanish":
      return 12
    case "SChinese":
      return 13
    case "TChinese":
      return 14
    case "Irish":
      return 15
  return -1

def stringToJson(in_files):
  stringData = {}
  for filename in in_files:
    if not os.path.isfile(filename):
      continue
    
    wb = load_workbook(filename, read_only = True)
    
    for s in wb:
      headers = []
      rows = s.iter_rows(min_col = 1, min_row = 1, max_col = 18, max_row = 1)
      for row in rows:
        for string in row[2:]:
          if string.value:
            headers.append(string.value)

      rows = s.iter_rows(min_col = 1, min_row = 1, max_col = 18, max_row = None)
      for row in rows:
        name = f"{row[0].value},{row[1].value}"
        if name == "Category,Id":
          continue
        if not name:
          continue
        
        data = {}
        
        for i, string in enumerate(row[2:]):
          if string.value:
            id = getLanguageId(headers[i])
            data[id] = string.value.replace("\r", "").replace("_x000D_", "").replace("\\n", "\n")
        
        if data:
          stringData[name] = data
  
  with open(OUT_FILE, "w", newline="\n") as f:
    json.dump(stringData, f, indent=4)

if __name__ == "__main__":
  in_files = [
    os.path.join(WORKING_DIR, "Strings.xlsx")
  ]
  
  stringToJson(in_files)